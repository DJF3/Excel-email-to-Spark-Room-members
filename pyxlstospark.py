
# ------------------ Add users from Excel file to Spark Room --------------------------
# AUTHOR: DJ (duittenb@cisco.com)
# DATE:   March 15th 2016
# VERSION:0.2
# More info: Github
# WARRANTY: Please consult your doctor before using this script

myToken="XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
myRoom="XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
myworkbook = "/Users/myname/Documents/SparkImport/attendees.xlsx"
mysheetStartCol = 3   # example: column A=1, B=2, C=3, D=4 ...
mysheetStartRow = 4   # row number
TestOnly = "yes"      # "yes"=only PRINT results, "no"=add users to Spark room

import json
import requests
import openpyxl
import re

def post_membership(mytoken,roomId,personEmail,isModerator=False):
    headers = {'Authorization':mytoken, 'content-type':'application/json'}
    payload = {'roomId':roomId, 'personEmail':personEmail, 'isModerator':isModerator}
    resp = requests.post(url='https://api.ciscospark.com/v1/memberships',json=payload, headers=headers)
    returnData = json.loads(resp.text)
    returnData['statuscode']=str(resp.status_code)
    return returnData

# ------------------------------------------------------------
# Read the spreadsheet, print the contents of one column
# ------------------------------------------------------------
ListToAdd = list()
ListSuccess = list()
ListFailed = list()
# __ Open Spreadsheet
wb = openpyxl.load_workbook(myworkbook)
# __ Getting name of FIRST sheet in workbook
mySheetName = wb.get_sheet_names()[0]
# __ Open the first sheet
sheet = wb.get_sheet_by_name(mySheetName)
# __ Get the highest row number (bottom of the sheet)
mysheethighestrow = sheet.max_row + 1

# __ Loop through all cells, identify correctly formatted
# __ email addresses and add these to a list
for i in range(mysheetStartRow, mysheethighestrow, 1):
    EmailFound = sheet.cell(row=i, column=mysheetStartCol).value
    if not re.match(r"[^@]+@[^@]+\.[^@]+", EmailFound):
        print "--- No valid email address found:", EmailFound
        ListFailed.append(EmailFound)
        continue
    if TestOnly.lower() == "yes": print EmailFound
    ListToAdd.append(EmailFound)

# ------------------------------------------------------------
# Add found email addresses to the Spark Room
# IF testOnly=NO: means PRODUCTION -> adding people to the room
# ------------------------------------------------------------
if TestOnly.lower() == "no" :
    # __ loop through found email addressess
    for addname in ListToAdd :
        print "====== ADD MEMBER:", addname
        # __ Add email address to configured Spark room
        txt2=post_membership("Bearer "+myToken,myRoom,addname)
        # __ Print results
        print "TXT: " + str(txt2)
        # __ Store email address in list if an error occurs
        if txt2.get('statuscode') == "403" :
            print "--- ERROR: status code is 403\n\n"
            ListFailed.append(addname)
        # __ Store email address in list if successfully added
        if txt2.get('statuscode') == "200" :
            ListSuccess.append(addname)

# __ Show failed email addressess and counts
print "\n\n ------------ Failed users: ------------ "
for names in ListFailed :
    print names
print "------------ failed count:", str(len(ListFailed))
print "------------ success count:", str(len(ListSuccess))

# -------------------------- the END ----------------------------------
